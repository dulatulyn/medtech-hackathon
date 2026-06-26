"""XLSX / XLS parser."""
from __future__ import annotations

import io
import subprocess
import tempfile
from pathlib import Path

import openpyxl

from src.enums import FileFormat
from src.parsers.base import ParseResult, RawRow, row_to_rawrow
from src.parsers.columns import find_header_row, map_columns


def _xls_to_xlsx(data: bytes) -> bytes:
    """Convert legacy .xls bytes to .xlsx using LibreOffice (soffice)."""
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "in.xls"
        src.write_bytes(data)
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(src)],
            check=True, capture_output=True, timeout=120,
        )
        out = next(Path(tmp).glob("*.xlsx"))
        return out.read_bytes()


def parse_excel(data: bytes, file_format: FileFormat) -> ParseResult:
    """Extract rows from every sheet of an Excel workbook."""
    if file_format == FileFormat.xls:
        data = _xls_to_xlsx(data)

    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[RawRow] = []
    warnings: list[str] = []
    for sheet_name in workbook.sheetnames:
        grid = [list(r) for r in workbook[sheet_name].iter_rows(values_only=True)]
        if not grid:
            continue
        header_idx = find_header_row(grid)
        cmap = map_columns(grid[header_idx])
        if not cmap.is_usable():
            warnings.append(f"sheet '{sheet_name}': no usable header")
            continue
        for row_idx in range(header_idx + 1, len(grid)):
            row = row_to_rawrow(grid[row_idx], cmap, {"sheet": sheet_name, "row": row_idx + 1})
            if row:
                rows.append(row)
    return ParseResult(rows, warnings)
