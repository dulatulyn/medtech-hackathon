"""Catalog service: load the reference dictionary and attach synonyms."""
from __future__ import annotations

import json
from pathlib import Path

from src.dtos.catalog_dto import DictionaryEntryDTO
from src.enums import SynonymSource
from src.repositories.catalog_repository import CatalogRepository


def _synonyms_from(value) -> list[str]:
    """Normalize a synonyms cell/field into a list of strings."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [s.strip() for s in str(value).replace(",", ";").split(";") if s.strip()]


def _entry_from_dict(d: dict) -> DictionaryEntryDTO:
    """Build a DictionaryEntryDTO from a loosely-keyed dict."""
    name = d.get("service_name") or d.get("name") or ""
    return DictionaryEntryDTO(
        name=str(name).strip(),
        category=(d.get("category") or None),
        icd_code=(d.get("icd_code") or d.get("code") or None),
        synonyms=_synonyms_from(d.get("synonyms")),
    )


def parse_dictionary_file(path: str) -> list[DictionaryEntryDTO]:
    """Parse a dictionary from JSON or XLSX into DTOs."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".json":
        data = json.loads(p.read_text(encoding="utf-8"))
        return [_entry_from_dict(d) for d in data if (d.get("service_name") or d.get("name"))]
    if suffix in (".xlsx", ".xls"):
        return _parse_xlsx(p)
    raise ValueError(f"unsupported dictionary format: {suffix}")


def _parse_xlsx(path: Path) -> list[DictionaryEntryDTO]:
    """Parse an XLSX dictionary; first row is the header."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    idx = {name: header.index(name) for name in header}

    def col(row, *names):
        for n in names:
            if n in idx and idx[n] < len(row):
                return row[idx[n]]
        return None

    entries: list[DictionaryEntryDTO] = []
    for row in rows:
        name = col(row, "service_name", "name", "наименование")
        if not name:
            continue
        entries.append(DictionaryEntryDTO(
            name=str(name).strip(),
            category=(col(row, "category", "категория") or None),
            icd_code=(col(row, "icd_code", "code", "код") or None),
            synonyms=_synonyms_from(col(row, "synonyms", "синонимы")),
        ))
    return entries


class CatalogService:
    """Business logic for the reference services dictionary."""

    def __init__(self, repo: CatalogRepository):
        self.repo = repo

    async def load_dictionary(self, entries: list[DictionaryEntryDTO]) -> int:
        """Upsert services and synonyms; idempotent. Returns services processed."""
        count = 0
        for entry in entries:
            if not entry.name:
                continue
            service = None
            if entry.icd_code:
                service = await self.repo.get_by_code(str(entry.icd_code))
            if service is None:
                service = await self.repo.get_by_name(entry.name)
            if service is None:
                service = await self.repo.create_service(entry.name, entry.category, entry.icd_code)
            else:
                if entry.category and not service.category:
                    service.category = entry.category
                if entry.icd_code and not service.icd_code:
                    service.icd_code = str(entry.icd_code)

            existing = {t.lower() for t in await self.repo.get_synonym_texts(service.id)}
            for syn in entry.synonyms:
                if syn.lower() not in existing:
                    await self.repo.add_synonym(service.id, syn, SynonymSource.dict)
                    existing.add(syn.lower())
            count += 1
        return count
